"""
Super Filter — SARI Organisation Statistics Generator
=======================================================
Reads a SARI survey Excel export and produces a multi-sheet workbook
matching the SARI_Organisation.xlsx format.

Sheets:
  1. Read Me              — Info
  2. Lists                — Organisation names
  3. Dashboard            — Interactive org dashboard (formulas)
  4. Organisation Report  — Per-org report (formulas)
  5. Organisation Summary — One row per org, stats
  6. Section Summary      — Per org × section stats
  7. Question Summary     — Per org × question stats
  8. Answer Distribution  — Per org × question × answer option
  9. Raw Answers          — Raw data with Standard section
  10. Priority Detail     — Priority-ranked questions per org

Usage:
    python super_filter.py
"""

import sys
import statistics
from pathlib import Path
from collections import defaultdict, Counter

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

# ═══════════════════════════════════════════════════════════════════════════
# CONFIG
# ═══════════════════════════════════════════════════════════════════════════

INPUT_FILE = "SARI_Results_2026-08-04-00-36-28.xlsx"
OUTPUT_FILE = "SARI_Organisation.xlsx"

RAW_COL_MAP = {
    1: "respondent_id", 2: "submitted_at", 3: "section",
    4: "question_num", 5: "question_id", 6: "question",
    7: "answer", 8: "answer_value", 9: "answer_score",
    10: "max_score", 11: "participant_name", 12: "email",
    13: "job_title", 14: "organisation_name", 15: "organisation_type",
    16: "organisation_size", 17: "stakeholder_category", 18: "pcds_sector",
    19: "district", 20: "role_level", 21: "department",
    22: "age_band", 23: "part_of_group", 24: "parent_company",
}

SECTION_ORDER = [
    "Background", "Strategy & Leadership", "Talent & Organisational Culture",
    "Data Management & Readiness", "Infrastructure & Technology",
    "Governance, Policy & Ethics", "Investment", "AI Implementation & Potential Impact",
]

BM_TO_EN_SECTION = {
    "Latar Belakang": "Background", "Strategi & Kepimpinan": "Strategy & Leadership",
    "Bakat & Budaya Organisasi": "Talent & Organisational Culture",
    "Pengurusan Data & Kesiapsiagaan": "Data Management & Readiness",
    "Infrastruktur & Teknologi": "Infrastructure & Technology",
    "Tadbir Urus, Dasar & Etika": "Governance, Policy & Ethics",
    "Pelaburan": "Investment", "Pelaksanaan AI & Impak": "AI Implementation & Potential Impact",
}

SCORED_SECTIONS = [
    "Strategy & Leadership", "Talent & Organisational Culture",
    "Data Management & Readiness", "Infrastructure & Technology",
    "Governance, Policy & Ethics", "Investment", "AI Implementation & Potential Impact",
]

MAX_SCORE = 4.0

# ═══════════════════════════════════════════════════════════════════════════
# STYLING
# ═══════════════════════════════════════════════════════════════════════════

HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="2F5496")
SUBTITLE_FONT = Font(name="Calibri", size=10, color="666666")
SECTION_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
SECTION_FONT = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
LIGHT_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_ALIGN = Alignment(vertical="top", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_header_row(ws, row, num_cols):
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def style_data_rows(ws, start_row, end_row, num_cols):
    for r in range(start_row, end_row + 1):
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER
            cell.alignment = DATA_ALIGN
            if (r - start_row) % 2 == 1:
                cell.fill = LIGHT_FILL


def auto_width(ws, num_cols, max_width=50, sample=200):
    for ci in range(1, num_cols + 1):
        best = 8
        for r in range(1, min(ws.max_row + 1, sample + 1)):
            v = str(ws.cell(r, ci).value or "")
            best = max(best, min(len(v), max_width))
        ws.column_dimensions[get_column_letter(ci)].width = min(best + 3, max_width)


# ═══════════════════════════════════════════════════════════════════════════
# DATA PROCESSING
# ═══════════════════════════════════════════════════════════════════════════

def read_raw(filepath):
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active
    rows = []
    for rc in ws.iter_rows(min_row=2, values_only=True):
        rec = {}
        for ci, key in RAW_COL_MAP.items():
            v = rc[ci - 1]
            rec[key] = str(v).strip() if v is not None else ""
        rows.append(rec)
    wb.close()
    return rows


def build_org_data(rows):
    orgs = defaultdict(lambda: {
        "type": "", "size": "", "sector": "", "district": "",
        "stakeholder": "", "part_of_group": "", "parent": "",
        "participants": set(), "departments": set(), "role_levels": set(),
        "age_bands": set(), "job_titles": set(), "latest_submission": "",
        "answers": defaultdict(list),  # (sec, qid) -> [(answer, score, answer_value), ...]
    })

    for row in rows:
        org = row["organisation_name"]
        if not org:
            continue
        o = orgs[org]

        if not o["type"]:
            o["type"] = row.get("organisation_type", "")
            o["size"] = row.get("organisation_size", "")
            o["sector"] = row.get("pcds_sector", "")
            o["district"] = row.get("district", "")
            o["stakeholder"] = row.get("stakeholder_category", "")
            o["part_of_group"] = row.get("part_of_group", "")
            o["parent"] = row.get("parent_company", "")

        o["participants"].add(row.get("participant_name", ""))
        if row.get("department"):
            o["departments"].add(row["department"])
        if row.get("role_level"):
            o["role_levels"].add(row["role_level"])
        if row.get("age_band"):
            o["age_bands"].add(row["age_band"])
        if row.get("job_title"):
            o["job_titles"].add(row["job_title"])

        submitted = row.get("submitted_at", "")
        if submitted and (not o["latest_submission"] or submitted > o["latest_submission"]):
            o["latest_submission"] = submitted

        sec = row["section"]
        qid = row["question_id"]
        en_sec = BM_TO_EN_SECTION.get(sec, sec)

        score_str = row.get("answer_score", "")
        try:
            score = float(score_str) if score_str else 0.0
        except ValueError:
            score = 0.0

        o["answers"][(en_sec, qid)].append({
            "answer": row.get("answer", ""),
            "score": score,
            "answer_value": row.get("answer_value", ""),
        })

    return orgs


def build_question_order(rows):
    seen = set()
    qs = []
    for row in rows:
        sec = row["section"]
        if sec not in SECTION_ORDER:
            continue
        qid = row["question_id"]
        key = (sec, qid)
        if key not in seen:
            seen.add(key)
            qs.append((sec, qid,
                int(row["question_num"]) if row["question_num"].isdigit() else 0,
                row["question"]))
    qs.sort(key=lambda x: (SECTION_ORDER.index(x[0]) if x[0] in SECTION_ORDER else 99, x[2]))
    return qs


# ═══════════════════════════════════════════════════════════════════════════
# METRICS COMPUTATION
# ═══════════════════════════════════════════════════════════════════════════

def consensus_score(answers_list):
    """Proportion of respondents giving the most common answer."""
    if not answers_list:
        return 0
    counts = Counter(a["answer"] for a in answers_list)
    return counts.most_common(1)[0][1] / len(answers_list)


def agreement_label(consensus, n_respondents):
    if n_respondents < 2:
        return "Not measurable"
    if consensus >= 0.8:
        return "High"
    if consensus >= 0.6:
        return "Moderate"
    return "Low"


def maturity_tier(overall):
    if overall is None:
        return ""
    if overall < 0.2:
        return "AI Aware - 0"
    if overall < 0.4:
        return "AI Explorer - 1"
    if overall < 0.6:
        return "AI Follower - 2"
    if overall < 0.8:
        return "AI Leader - 3"
    return "AI Pioneer - 4"


def distance_to_next_tier(overall):
    if overall is None:
        return ""
    if overall >= 0.8:
        return round(1.0 - overall, 4)
    return round((((overall // 0.2) + 1) * 0.2) - overall, 4)


def compute_org_summary(orgs, questions):
    """Compute Organisation Summary data."""
    rows = []
    for org_name in sorted(orgs.keys()):
        o = orgs[org_name]
        n_respondents = len(o["participants"])

        # Per-question stats
        q_stats = {}
        all_scores = []
        for sec, qid, qnum, qtext in questions:
            ans_list = o["answers"].get((sec, qid), [])
            scores = [a["score"] for a in ans_list]
            if scores:
                q_stats[(sec, qid)] = {
                    "n": len(ans_list),
                    "scores": scores,
                    "avg": sum(scores) / len(scores),
                    "consensus": consensus_score(ans_list),
                    "most_common": Counter(a["answer"] for a in ans_list).most_common(1)[0][0] if ans_list else "",
                }
                all_scores.extend(scores)

        # Average score (raw 0-4)
        avg_score = sum(s["avg"] for s in q_stats.values()) / len(q_stats) if q_stats else 0
        # Overall score (normalized 0-1)
        overall = avg_score / MAX_SCORE if MAX_SCORE > 0 else 0

        # Per-section normalized scores
        sec_scores = {}
        for sec in SCORED_SECTIONS:
            sec_qs = [(sq, sqid) for sq, sqid, _, _ in questions if sq == sec]
            sec_avgs = [q_stats[(sq, sqid)]["avg"] for sq, sqid in sec_qs if (sq, sqid) in q_stats]
            if sec_avgs:
                sec_scores[sec] = sum(sec_avgs) / len(sec_avgs) / MAX_SCORE

        strongest = max(sec_scores, key=sec_scores.get) if sec_scores else ""
        weakest = min(sec_scores, key=sec_scores.get) if sec_scores else ""

        # Average consensus
        consensuses = [s["consensus"] for s in q_stats.values()]
        avg_consensus = sum(consensuses) / len(consensuses) if consensuses else 0

        # Questions for review (consensus < 0.6 or high dispersion)
        review_count = sum(1 for s in q_stats.values() if s["consensus"] < 0.6)

        # Agreement
        agreement = agreement_label(avg_consensus, n_respondents)

        # Interpretation
        interpretation = "Multi-respondent view" if n_respondents > 1 else "Single-respondent view"

        rows.append({
            "org_name": org_name,
            "org_type": o["type"],
            "respondents": n_respondents,
            "departments": len(o["departments"]),
            "role_levels": len(o["role_levels"]),
            "org_size": o["size"],
            "sector": o["sector"],
            "latest_submission": o["latest_submission"],
            "avg_score": round(avg_score, 4),
            "overall_score": round(overall, 4),
            "strongest": strongest,
            "weakest": weakest,
            "avg_consensus": round(avg_consensus, 4),
            "review_count": review_count,
            "agreement": agreement,
            "interpretation": interpretation,
            "maturity_tier": maturity_tier(overall),
            "distance": distance_to_next_tier(overall),
        })
    return rows


def compute_section_summary(orgs, questions):
    """Per org × section stats."""
    rows = []
    for org_name in sorted(orgs.keys()):
        o = orgs[org_name]
        for sec in SCORED_SECTIONS:
            sec_qs = [(sq, sqid, sqnum, sqtext) for sq, sqid, sqnum, sqtext in questions if sq == sec]
            all_scores = []
            all_consensus = []
            n_respondents_set = set()
            for sq, sqid, _, _ in sec_qs:
                ans_list = o["answers"].get((sq, sqid), [])
                if ans_list:
                    scores = [a["score"] for a in ans_list]
                    all_scores.extend(scores)
                    all_consensus.append(consensus_score(ans_list))
                    n_respondents_set.add(len(ans_list))

            if all_scores:
                avg = sum(all_scores) / len(all_scores)
                med = statistics.median(all_scores)
                mn = min(all_scores)
                mx = max(all_scores)
                norm = avg / MAX_SCORE
                avg_con = sum(all_consensus) / len(all_consensus) if all_consensus else 0
                n_resp = max(n_respondents_set) if n_respondents_set else 0
                ag = agreement_label(avg_con, n_resp)
            else:
                avg = med = mn = mx = norm = avg_con = 0
                n_resp = 0
                ag = "Not measurable"

            rows.append({
                "org_name": org_name,
                "section": sec,
                "respondents": n_resp,
                "questions": len(sec_qs),
                "avg_score": round(avg, 4),
                "median_score": round(med, 4),
                "min_score": round(mn, 4),
                "max_score": round(mx, 4),
                "max_possible": MAX_SCORE,
                "normalised": round(norm, 4),
                "avg_consensus": round(avg_con, 4),
                "agreement": ag,
            })
    return rows


def compute_question_summary(orgs, questions):
    """Per org × question stats."""
    rows = []
    for org_name in sorted(orgs.keys()):
        o = orgs[org_name]
        for sec, qid, qnum, qtext in questions:
            ans_list = o["answers"].get((sec, qid), [])
            n = len(ans_list)
            scores = [a["score"] for a in ans_list]
            is_scored = sec in SCORED_SECTIONS

            if scores:
                avg = sum(scores) / len(scores)
                med = statistics.median(scores)
                mn = min(scores)
                mx = max(scores)
                score_range = mx - mn
                std = statistics.stdev(scores) if len(scores) > 1 else 0
                norm = avg / MAX_SCORE
                con = consensus_score(ans_list)
                most_common = Counter(a["answer"] for a in ans_list).most_common(1)
                mc_answer = most_common[0][0] if most_common else ""
                mc_count = most_common[0][1] if most_common else 0
                ag = agreement_label(con, n)
                # Review flag: low consensus or high std dev
                review = "REVIEW" if (con < 0.6 or std > 1.5) else None
            else:
                avg = med = mn = mx = score_range = std = norm = con = mc_count = 0
                mc_answer = ""
                ag = "Not measurable"
                review = None

            rows.append({
                "org_name": org_name,
                "section": sec,
                "qid": qid,
                "question": qtext,
                "respondents": n,
                "scored": is_scored,
                "most_common_answer": mc_answer,
                "most_common_count": mc_count,
                "consensus": round(con, 4),
                "avg_score": round(avg, 4),
                "median_score": round(med, 4),
                "min_score": round(mn, 4),
                "max_score": round(mx, 4),
                "score_range": round(score_range, 4),
                "std_dev": round(std, 4),
                "normalised": round(norm, 4),
                "agreement": ag,
                "review_flag": review,
            })
    return rows


def compute_answer_distribution(orgs, questions):
    """Per org × question × answer option counts."""
    rows = []
    for org_name in sorted(orgs.keys()):
        o = orgs[org_name]
        for sec, qid, qnum, qtext in questions:
            ans_list = o["answers"].get((sec, qid), [])
            n_total = len(ans_list)

            # Determine question type
            is_multi = any("|" in a["answer_value"] for a in ans_list)
            qtype = "Multi-select" if is_multi else "Single-choice"

            # Count by answer option
            answer_counts = Counter()
            answer_scores = {}
            for a in ans_list:
                answer_counts[a["answer"]] += 1
                answer_scores[a["answer"]] = a["score"]

            for answer_text, count in answer_counts.items():
                rows.append({
                    "org_name": org_name,
                    "section": sec,
                    "qid": qid,
                    "question": qtext,
                    "qtype": qtype,
                    "answer_option": answer_text,
                    "answer_score": answer_scores.get(answer_text, 0),
                    "respondents_selecting": count,
                    "percentage": round(count / n_total, 4) if n_total > 0 else 0,
                    "question_respondents": n_total,
                })
    return rows


def compute_priority_detail(orgs, questions):
    """Priority-ranked questions per org (low score + low consensus first)."""
    rows = []
    for org_name in sorted(orgs.keys()):
        o = orgs[org_name]
        org_priorities = []
        for sec, qid, qnum, qtext in questions:
            if sec not in SCORED_SECTIONS:
                continue
            ans_list = o["answers"].get((sec, qid), [])
            if not ans_list:
                continue
            scores = [a["score"] for a in ans_list]
            avg = sum(scores) / len(scores)
            norm = avg / MAX_SCORE
            con = consensus_score(ans_list)
            most_common = Counter(a["answer"] for a in ans_list).most_common(1)[0][0]
            ag = agreement_label(con, len(ans_list))
            review = "REVIEW" if (con < 0.6 or (len(scores) > 1 and statistics.stdev(scores) > 1.5)) else None

            # Priority: sort by normalized score ascending, then consensus ascending
            org_priorities.append({
                "qid": qid,
                "question": qtext,
                "most_common": most_common,
                "normalised": round(norm, 4),
                "agreement": ag,
                "review_flag": review,
                "sort_key": (norm, con),
            })

        org_priorities.sort(key=lambda x: x["sort_key"])
        for rank, p in enumerate(org_priorities, 1):
            rows.append({
                "org_name": org_name,
                "qid": p["qid"],
                "question": p["question"],
                "most_common": p["most_common"],
                "normalised": p["normalised"],
                "agreement": p["agreement"],
                "review_flag": p["review_flag"],
                "priority_rank": rank,
                "lookup_key": f"{org_name}|{rank}",
            })
    return rows


# ═══════════════════════════════════════════════════════════════════════════
# SHEET BUILDERS
# ═══════════════════════════════════════════════════════════════════════════

def write_read_me(wb):
    ws = wb.create_sheet("Read Me", 0)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 100
    ws.cell(1, 1).value = "SARI Organisation Statistics"
    ws.cell(1, 1).font = TITLE_FONT
    ws.cell(2, 1).value = "Interactive organisation-level reporting workbook generated from the uploaded SARI results."
    ws.cell(2, 1).font = SUBTITLE_FONT
    ws.cell(4, 1).value = "Purpose"
    ws.cell(4, 1).font = Font(bold=True)
    ws.cell(4, 2).value = "Analyse questionnaire results by organisation, section, question and answer distribution."


def write_lists(wb, orgs):
    ws = wb.create_sheet("Lists")
    for i, name in enumerate(sorted(orgs.keys()), 1):
        ws.cell(i, 1).value = name


def write_org_summary_sheet(wb, orgs, questions):
    ws = wb.create_sheet("Organisation Summary")
    headers = [
        "Organisation name", "Organisation type", "Respondents",
        "Departments represented", "Role levels represented",
        "Organisation size", "Sector", "Latest submission",
        "Average score", "Overall score", "Strongest section",
        "Weakest section", "Average consensus", "Questions for review",
        "Agreement", "Interpretation", "Maturity tier", "Distance to next tier",
    ]

    ws.cell(1, 1).value = "Organisation Summary"
    ws.cell(1, 1).font = TITLE_FONT
    ws.cell(2, 1).value = "One row per organisation. Sort or filter by respondent count, score or agreement."
    ws.cell(2, 1).font = SUBTITLE_FONT

    for ci, h in enumerate(headers, 1):
        ws.cell(4, ci).value = h
    style_header_row(ws, 4, len(headers))

    data = compute_org_summary(orgs, questions)
    for ri, d in enumerate(data, 5):
        ws.cell(ri, 1).value = d["org_name"]
        ws.cell(ri, 2).value = d["org_type"]
        ws.cell(ri, 3).value = d["respondents"]
        ws.cell(ri, 4).value = d["departments"]
        ws.cell(ri, 5).value = d["role_levels"]
        ws.cell(ri, 6).value = d["org_size"]
        ws.cell(ri, 7).value = d["sector"]
        ws.cell(ri, 8).value = d["latest_submission"]
        ws.cell(ri, 9).value = d["avg_score"]
        ws.cell(ri, 9).number_format = '0.0000'
        ws.cell(ri, 10).value = d["overall_score"]
        ws.cell(ri, 10).number_format = '0.0000'
        ws.cell(ri, 11).value = d["strongest"]
        ws.cell(ri, 12).value = d["weakest"]
        ws.cell(ri, 13).value = d["avg_consensus"]
        ws.cell(ri, 13).number_format = '0.0000'
        ws.cell(ri, 14).value = d["review_count"]
        ws.cell(ri, 15).value = d["agreement"]
        ws.cell(ri, 16).value = d["interpretation"]
        # Maturity tier: formula for interactivity
        ws.cell(ri, 17).value = f'=IF(J{ri}="","",IF(J{ri}<0.2,"AI Aware - 0",IF(J{ri}<0.4,"AI Explorer - 1",IF(J{ri}<0.6,"AI Follower - 2",IF(J{ri}<0.8,"AI Leader - 3","AI Pioneer - 4")))))'
        # Distance to next tier: formula
        ws.cell(ri, 18).value = f'=IF(J{ri}="","",IF(J{ri}>=0.8,1-J{ri},CEILING(J{ri},0.2)-J{ri}))'

    last_row = 4 + len(data)
    style_data_rows(ws, 5, last_row, len(headers))
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{last_row}"
    auto_width(ws, len(headers))

    # Color scale on score columns
    score_range = f"I5:J{last_row}"
    ws.conditional_formatting.add(score_range, ColorScaleRule(
        start_type="min", start_color="F8696B",
        mid_type="percentile", mid_value=50, mid_color="FFEB84",
        end_type="max", end_color="63BE7B",
    ))

    return data


def write_section_summary_sheet(wb, orgs, questions):
    ws = wb.create_sheet("Section Summary")
    headers = [
        "Organisation name", "Section", "Respondents", "Questions",
        "Average score", "Median score", "Minimum score", "Maximum score",
        "Max possible", "Normalised score", "Average consensus", "Agreement",
    ]

    ws.cell(1, 1).value = "Section Summary"
    ws.cell(1, 1).font = TITLE_FONT
    ws.cell(2, 1).value = "Section-level maturity and internal agreement for each organisation."
    ws.cell(2, 1).font = SUBTITLE_FONT

    for ci, h in enumerate(headers, 1):
        ws.cell(4, ci).value = h
    style_header_row(ws, 4, len(headers))

    data = compute_section_summary(orgs, questions)
    for ri, d in enumerate(data, 5):
        ws.cell(ri, 1).value = d["org_name"]
        ws.cell(ri, 2).value = d["section"]
        ws.cell(ri, 3).value = d["respondents"]
        ws.cell(ri, 4).value = d["questions"]
        ws.cell(ri, 5).value = d["avg_score"]
        ws.cell(ri, 5).number_format = '0.0000'
        ws.cell(ri, 6).value = d["median_score"]
        ws.cell(ri, 6).number_format = '0.0000'
        ws.cell(ri, 7).value = d["min_score"]
        ws.cell(ri, 7).number_format = '0.0000'
        ws.cell(ri, 8).value = d["max_score"]
        ws.cell(ri, 8).number_format = '0.0000'
        ws.cell(ri, 9).value = d["max_possible"]
        ws.cell(ri, 10).value = d["normalised"]
        ws.cell(ri, 10).number_format = '0.0000'
        ws.cell(ri, 11).value = d["avg_consensus"]
        ws.cell(ri, 11).number_format = '0.0000'
        ws.cell(ri, 12).value = d["agreement"]

    last_row = 4 + len(data)
    style_data_rows(ws, 5, last_row, len(headers))
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{last_row}"
    auto_width(ws, len(headers))

    # Color scale
    score_range = f"E5:J{last_row}"
    ws.conditional_formatting.add(score_range, ColorScaleRule(
        start_type="min", start_color="F8696B",
        mid_type="percentile", mid_value=50, mid_color="FFEB84",
        end_type="max", end_color="63BE7B",
    ))

    return data


def write_question_summary_sheet(wb, orgs, questions):
    ws = wb.create_sheet("Question Summary")
    headers = [
        "Organisation name", "Section", "Question ID", "Question",
        "Respondents", "Scored question", "Most common answer",
        "Most common count", "Consensus", "Average score",
        "Median score", "Minimum score", "Maximum score",
        "Score range", "Standard deviation", "Normalised score",
        "Agreement", "Review flag",
    ]

    ws.cell(1, 1).value = "Question Summary"
    ws.cell(1, 1).font = TITLE_FONT
    ws.cell(2, 1).value = "Question-level statistics. Use Review flag to find low-consensus or highly dispersed scored questions."
    ws.cell(2, 1).font = SUBTITLE_FONT

    for ci, h in enumerate(headers, 1):
        ws.cell(4, ci).value = h
    style_header_row(ws, 4, len(headers))

    data = compute_question_summary(orgs, questions)
    for ri, d in enumerate(data, 5):
        ws.cell(ri, 1).value = d["org_name"]
        ws.cell(ri, 2).value = d["section"]
        ws.cell(ri, 3).value = d["qid"]
        ws.cell(ri, 4).value = d["question"]
        ws.cell(ri, 5).value = d["respondents"]
        ws.cell(ri, 6).value = d["scored"]
        ws.cell(ri, 7).value = d["most_common_answer"]
        ws.cell(ri, 8).value = d["most_common_count"]
        ws.cell(ri, 9).value = d["consensus"]
        ws.cell(ri, 9).number_format = '0.0000'
        ws.cell(ri, 10).value = d["avg_score"]
        ws.cell(ri, 10).number_format = '0.0000'
        ws.cell(ri, 11).value = d["median_score"]
        ws.cell(ri, 11).number_format = '0.0000'
        ws.cell(ri, 12).value = d["min_score"]
        ws.cell(ri, 12).number_format = '0.0000'
        ws.cell(ri, 13).value = d["max_score"]
        ws.cell(ri, 13).number_format = '0.0000'
        ws.cell(ri, 14).value = d["score_range"]
        ws.cell(ri, 14).number_format = '0.0000'
        ws.cell(ri, 15).value = d["std_dev"]
        ws.cell(ri, 15).number_format = '0.0000'
        ws.cell(ri, 16).value = d["normalised"]
        ws.cell(ri, 16).number_format = '0.0000'
        ws.cell(ri, 17).value = d["agreement"]
        ws.cell(ri, 18).value = d["review_flag"]

    last_row = 4 + len(data)
    style_data_rows(ws, 5, last_row, len(headers))
    ws.freeze_panes = "D5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{last_row}"
    auto_width(ws, len(headers))

    return data


def write_answer_distribution_sheet(wb, orgs, questions):
    ws = wb.create_sheet("Answer Distribution")
    headers = [
        "Organisation name", "Standard section", "Question ID", "Question",
        "Question type", "Answer option", "Answer score",
        "Respondents selecting", "Percentage", "Question respondents",
    ]

    ws.cell(1, 1).value = "Answer Distribution"
    ws.cell(1, 1).font = TITLE_FONT
    ws.cell(2, 1).value = "Counts and percentages by answer option. Multi-select totals can exceed 100%."
    ws.cell(2, 1).font = SUBTITLE_FONT

    for ci, h in enumerate(headers, 1):
        ws.cell(4, ci).value = h
    style_header_row(ws, 4, len(headers))

    data = compute_answer_distribution(orgs, questions)
    for ri, d in enumerate(data, 5):
        ws.cell(ri, 1).value = d["org_name"]
        ws.cell(ri, 2).value = d["section"]
        ws.cell(ri, 3).value = d["qid"]
        ws.cell(ri, 4).value = d["question"]
        ws.cell(ri, 5).value = d["qtype"]
        ws.cell(ri, 6).value = d["answer_option"]
        ws.cell(ri, 7).value = d["answer_score"]
        ws.cell(ri, 8).value = d["respondents_selecting"]
        ws.cell(ri, 9).value = d["percentage"]
        ws.cell(ri, 9).number_format = '0.00%'
        ws.cell(ri, 10).value = d["question_respondents"]

    last_row = 4 + len(data)
    style_data_rows(ws, 5, last_row, len(headers))
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{last_row}"
    auto_width(ws, len(headers))

    return data


def write_raw_answers_sheet(wb, rows):
    ws = wb.create_sheet("Raw Answers")
    headers = [
        "Respondent ID", "Submitted at", "Standard section", "Question #",
        "Question ID", "Question", "Answer", "Answer value",
        "Answer score", "Max score", "Participant name", "Job title",
        "Organisation name", "Organisation type", "Organisation size",
        "Stakeholder category", "PCDS sector", "District",
        "Role level", "Department", "Age band", "Part of group",
        "Parent company",
    ]

    ws.cell(2, 1).value = "Imported source rows with an added Standard section field. Personal email is intentionally omitted from this analysis copy."
    ws.cell(2, 1).font = SUBTITLE_FONT

    for ci, h in enumerate(headers, 1):
        ws.cell(4, ci).value = h
    style_header_row(ws, 4, len(headers))

    for ri, row in enumerate(rows, 5):
        ws.cell(ri, 1).value = row.get("respondent_id", "")
        ws.cell(ri, 2).value = row.get("submitted_at", "")
        # Standard section = BM→EN merged
        sec = row.get("section", "")
        ws.cell(ri, 3).value = BM_TO_EN_SECTION.get(sec, sec)
        ws.cell(ri, 4).value = int(row["question_num"]) if row.get("question_num", "").isdigit() else row.get("question_num", "")
        ws.cell(ri, 5).value = row.get("question_id", "")
        ws.cell(ri, 6).value = row.get("question", "")
        ws.cell(ri, 7).value = row.get("answer", "")
        ws.cell(ri, 8).value = row.get("answer_value", "")
        try:
            ws.cell(ri, 9).value = float(row.get("answer_score", 0))
        except ValueError:
            ws.cell(ri, 9).value = 0
        try:
            ws.cell(ri, 10).value = float(row.get("max_score", 0)) if row.get("max_score") else None
        except ValueError:
            ws.cell(ri, 10).value = None
        ws.cell(ri, 11).value = row.get("participant_name", "")
        ws.cell(ri, 12).value = row.get("job_title", "")
        ws.cell(ri, 13).value = row.get("organisation_name", "")
        ws.cell(ri, 14).value = row.get("organisation_type", "")
        ws.cell(ri, 15).value = row.get("organisation_size", "")
        ws.cell(ri, 16).value = row.get("stakeholder_category", "")
        ws.cell(ri, 17).value = row.get("pcds_sector", "")
        ws.cell(ri, 18).value = row.get("district", "")
        ws.cell(ri, 19).value = row.get("role_level", "")
        ws.cell(ri, 20).value = row.get("department", "")
        ws.cell(ri, 21).value = row.get("age_band", "")
        ws.cell(ri, 22).value = row.get("part_of_group", "")
        ws.cell(ri, 23).value = row.get("parent_company", "")

    last_row = 4 + len(rows)
    style_data_rows(ws, 5, last_row, len(headers))
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{last_row}"
    auto_width(ws, len(headers))


def write_priority_detail_sheet(wb, orgs, questions):
    ws = wb.create_sheet("Priority Detail")
    headers = [
        "Organisation", "Question ID", "Question", "Most common answer",
        "Normalised score", "Agreement", "Review flag", "Priority rank", "Lookup key",
    ]

    for ci, h in enumerate(headers, 1):
        ws.cell(1, ci).value = h
    style_header_row(ws, 1, len(headers))

    data = compute_priority_detail(orgs, questions)
    for ri, d in enumerate(data, 2):
        ws.cell(ri, 1).value = d["org_name"]
        ws.cell(ri, 2).value = d["qid"]
        ws.cell(ri, 3).value = d["question"]
        ws.cell(ri, 4).value = d["most_common"]
        ws.cell(ri, 5).value = d["normalised"]
        ws.cell(ri, 5).number_format = '0.0000'
        ws.cell(ri, 6).value = d["agreement"]
        ws.cell(ri, 7).value = d["review_flag"]
        ws.cell(ri, 8).value = d["priority_rank"]
        ws.cell(ri, 9).value = d["lookup_key"]

    last_row = 1 + len(data)
    style_data_rows(ws, 2, last_row, len(headers))
    auto_width(ws, len(headers))


def write_dashboard_sheet(wb, orgs):
    ws = wb.create_sheet("Dashboard")
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 39
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["G"].width = 31

    ws.cell(1, 1).value = "Organisation Dashboard"
    ws.cell(1, 1).font = TITLE_FONT
    ws.cell(2, 1).value = "Select an organisation. Green cells contain imported selections; black cells are formulas."
    ws.cell(2, 1).font = SUBTITLE_FONT

    # Selected organisation (green cell)
    ws.cell(4, 1).value = "Selected organisation"
    ws.cell(4, 1).font = Font(bold=True)
    first_org = sorted(orgs.keys())[0] if orgs else ""
    ws.cell(4, 2).value = first_org
    ws.cell(4, 2).fill = GREEN_FILL
    ws.cell(4, 2).font = Font(bold=True)

    # Stats with formulas
    stats = [
        (6, "Respondents", 7, "Overall score", 6, "Strongest section"),
        (9, "Weakest section", 10, "Agreement", 9, "Interpretation"),
        (12, "Maturity tier", 13, "Distance to next tier", 12, "Questions for review"),
    ]
    for label_row, label1, val_row, label2, label_row2, label3 in stats:
        ws.cell(label_row, 1).value = label1
        ws.cell(label_row, 1).font = Font(bold=True)
        ws.cell(label_row, 4).value = label2
        ws.cell(label_row, 4).font = Font(bold=True)
        ws.cell(label_row, 7).value = label3
        ws.cell(label_row, 7).font = Font(bold=True)

        ws.cell(val_row, 1).value = f'=IFERROR(INDEX(\'Organisation Summary\'!$C:$C,MATCH($B$4,\'Organisation Summary\'!$A:$A,0)),"")'
        ws.cell(val_row, 4).value = f'=IFERROR(INDEX(\'Organisation Summary\'!$J:$J,MATCH($B$4,\'Organisation Summary\'!$A:$A,0)),"")'
        ws.cell(val_row, 7).value = f'=IFERROR(INDEX(\'Organisation Summary\'!$K:$K,MATCH($B$4,\'Organisation Summary\'!$A:$A,0)),"")'

    # Weakest/Agreement/Interpretation row
    ws.cell(10, 1).value = f'=IFERROR(INDEX(\'Organisation Summary\'!$L:$L,MATCH($B$4,\'Organisation Summary\'!$A:$A,0)),"")'
    ws.cell(10, 4).value = f'=IFERROR(INDEX(\'Organisation Summary\'!$O:$O,MATCH($B$4,\'Organisation Summary\'!$A:$A,0)),"")'
    ws.cell(10, 7).value = f'=IFERROR(INDEX(\'Organisation Summary\'!$P:$P,MATCH($B$4,\'Organisation Summary\'!$A:$A,0)),"")'

    # Maturity/Distance/Review row
    ws.cell(13, 1).value = f'=IFERROR(INDEX(\'Organisation Summary\'!$Q:$Q,MATCH($B$4,\'Organisation Summary\'!$A:$A,0)),"")'
    ws.cell(13, 4).value = f'=IFERROR(INDEX(\'Organisation Summary\'!$R:$R,MATCH($B$4,\'Organisation Summary\'!$A:$A,0)),"")'
    ws.cell(13, 7).value = f'=IFERROR(INDEX(\'Organisation Summary\'!$N:$N,MATCH($B$4,\'Organisation Summary\'!$A:$A,0)),"")'

    # Section breakdown
    ws.cell(17, 1).value = "Section"
    ws.cell(17, 1).font = Font(bold=True)
    ws.cell(17, 2).value = "Average score"
    ws.cell(17, 2).font = Font(bold=True)
    ws.cell(17, 3).value = "Normalised score"
    ws.cell(17, 3).font = Font(bold=True)
    ws.cell(17, 4).value = "Respondents"
    ws.cell(17, 4).font = Font(bold=True)
    ws.cell(17, 5).value = "Agreement"
    ws.cell(17, 5).font = Font(bold=True)

    for i, sec in enumerate(SCORED_SECTIONS):
        r = 18 + i
        ws.cell(r, 1).value = sec
        ws.cell(r, 2).value = f'=IFERROR(SUMIFS(\'Section Summary\'!$E:$E,\'Section Summary\'!$A:$A,$B$4,\'Section Summary\'!$B:$B,$A{r}),"")'
        ws.cell(r, 2).number_format = '0.0000'
        ws.cell(r, 3).value = f'=IFERROR(SUMIFS(\'Section Summary\'!$J:$J,\'Section Summary\'!$A:$A,$B$4,\'Section Summary\'!$B:$B,$A{r}),"")'
        ws.cell(r, 3).number_format = '0.0000'
        ws.cell(r, 4).value = f'=IFERROR(SUMIFS(\'Section Summary\'!$C:$C,\'Section Summary\'!$A:$A,$B$4,\'Section Summary\'!$B:$B,$A{r}),"")'
        ws.cell(r, 5).value = f'=IF(D{r}<2,"Not measurable",IF(SUMIFS(\'Section Summary\'!$K:$K,\'Section Summary\'!$A:$A,$B$4,\'Section Summary\'!$B:$B,$A{r})>=0.8,"High",IF(SUMIFS(\'Section Summary\'!$K:$K,\'Section Summary\'!$A:$A,$B$4,\'Section Summary\'!$B:$B,$A{r})>=0.6,"Moderate","Low")))'

    ws.freeze_panes = "A17"


def write_org_report_sheet(wb, orgs):
    ws = wb.create_sheet("Organisation Report")
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 18

    ws.cell(1, 1).value = "SARI Organisation Report"
    ws.cell(1, 1).font = TITLE_FONT

    ws.cell(3, 1).value = "Selected organisation"
    ws.cell(3, 1).font = Font(bold=True)
    first_org = sorted(orgs.keys())[0] if orgs else ""
    ws.cell(3, 2).value = first_org
    ws.cell(3, 2).fill = GREEN_FILL
    ws.cell(3, 2).font = Font(bold=True)

    # Key stats
    ws.cell(5, 1).value = "Overall score"
    ws.cell(5, 1).font = Font(bold=True)
    ws.cell(5, 2).value = f'=IFERROR(INDEX(\'Organisation Summary\'!$J:$J,MATCH($B$3,\'Organisation Summary\'!$A:$A,0)),"")'
    ws.cell(5, 2).number_format = '0.0000'
    ws.cell(5, 4).value = "Maturity tier"
    ws.cell(5, 4).font = Font(bold=True)
    ws.cell(5, 5).value = f'=IFERROR(INDEX(\'Organisation Summary\'!$Q:$Q,MATCH($B$3,\'Organisation Summary\'!$A:$A,0)),"")'
    ws.cell(5, 7).value = "Respondents"
    ws.cell(5, 7).font = Font(bold=True)
    ws.cell(5, 8).value = f'=IFERROR(INDEX(\'Organisation Summary\'!$C:$C,MATCH($B$3,\'Organisation Summary\'!$A:$A,0)),"")'

    ws.cell(7, 1).value = "Strongest section"
    ws.cell(7, 1).font = Font(bold=True)
    ws.cell(7, 2).value = f'=IFERROR(INDEX(\'Organisation Summary\'!$K:$K,MATCH($B$3,\'Organisation Summary\'!$A:$A,0)),"")'
    ws.cell(7, 4).value = "Weakest section"
    ws.cell(7, 4).font = Font(bold=True)
    ws.cell(7, 5).value = f'=IFERROR(INDEX(\'Organisation Summary\'!$L:$L,MATCH($B$3,\'Organisation Summary\'!$A:$A,0)),"")'
    ws.cell(7, 7).value = "Agreement"
    ws.cell(7, 7).font = Font(bold=True)
    ws.cell(7, 8).value = f'=IFERROR(INDEX(\'Organisation Summary\'!$O:$O,MATCH($B$3,\'Organisation Summary\'!$A:$A,0)),"")'

    # Section scores
    ws.cell(10, 1).value = "Section"
    ws.cell(10, 1).font = Font(bold=True)
    ws.cell(10, 2).value = "Score"
    ws.cell(10, 2).font = Font(bold=True)
    ws.cell(10, 3).value = "Agreement"
    ws.cell(10, 3).font = Font(bold=True)

    for i, sec in enumerate(SCORED_SECTIONS):
        r = 11 + i
        ws.cell(r, 1).value = sec
        ws.cell(r, 2).value = f'=IFERROR(SUMIFS(\'Section Summary\'!$J:$J,\'Section Summary\'!$A:$A,$B$3,\'Section Summary\'!$B:$B,$A{r}),"")'
        ws.cell(r, 2).number_format = '0.0000'
        ws.cell(r, 3).value = f'=IFERROR(LOOKUP(2,1/((\'Section Summary\'!$A$5:$A$872=$B$3)*(\'Section Summary\'!$B$5:$B$872=$A{r})),\'Section Summary\'!$L$5:$L$872),"")'

    # Priority questions
    ws.cell(20, 1).value = "Priority questions for review"
    ws.cell(20, 1).font = Font(bold=True)
    ws.cell(21, 1).value = "Question ID"
    ws.cell(21, 1).font = Font(bold=True)
    ws.cell(21, 2).value = "Question"
    ws.cell(21, 2).font = Font(bold=True)
    ws.cell(21, 3).value = "Most common answer"
    ws.cell(21, 3).font = Font(bold=True)
    ws.cell(21, 4).value = "Normalised score"
    ws.cell(21, 4).font = Font(bold=True)
    ws.cell(21, 5).value = "Agreement"
    ws.cell(21, 5).font = Font(bold=True)
    ws.cell(21, 6).value = "Review flag"
    ws.cell(21, 6).font = Font(bold=True)

    for rank in range(1, 6):
        r = 21 + rank
        ws.cell(r, 1).value = f'=IFERROR(INDEX(\'Priority Detail\'!$B:$B,MATCH($B$3&"|{rank}",\'Priority Detail\'!$I:$I,0)),"")'
        ws.cell(r, 2).value = f'=IFERROR(INDEX(\'Priority Detail\'!$C:$C,MATCH($B$3&"|{rank}",\'Priority Detail\'!$I:$I,0)),"")'
        ws.cell(r, 3).value = f'=IFERROR(INDEX(\'Priority Detail\'!$D:$D,MATCH($B$3&"|{rank}",\'Priority Detail\'!$I:$I,0)),"")'
        ws.cell(r, 4).value = f'=IFERROR(INDEX(\'Priority Detail\'!$E:$E,MATCH($B$3&"|{rank}",\'Priority Detail\'!$I:$I,0)),"")'
        ws.cell(r, 5).value = f'=IFERROR(INDEX(\'Priority Detail\'!$F:$F,MATCH($B$3&"|{rank}",\'Priority Detail\'!$I:$I,0)),"")'
        ws.cell(r, 6).value = f'=IFERROR(INDEX(\'Priority Detail\'!$G:$G,MATCH($B$3&"|{rank}",\'Priority Detail\'!$I:$I,0)),"")'

    ws.freeze_panes = "A10"


# ═══════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════

def main():
    input_path = Path(INPUT_FILE)
    if not input_path.exists():
        print(f"ERROR: Input file not found: {INPUT_FILE}")
        sys.exit(1)

    print(f"Reading: {INPUT_FILE}")
    rows = read_raw(str(input_path))
    print(f"  {len(rows)} raw rows")

    questions = build_question_order(rows)
    print(f"  {len(questions)} unique questions")

    orgs = build_org_data(rows)
    print(f"  {len(orgs)} organisations")

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    print("Writing sheets...")
    write_read_me(wb)
    print("  1. Read Me")
    write_lists(wb, orgs)
    print("  2. Lists")
    write_org_summary_sheet(wb, orgs, questions)
    print("  3. Organisation Summary")
    write_section_summary_sheet(wb, orgs, questions)
    print("  4. Section Summary")
    write_question_summary_sheet(wb, orgs, questions)
    print("  5. Question Summary")
    write_answer_distribution_sheet(wb, orgs, questions)
    print("  6. Answer Distribution")
    write_raw_answers_sheet(wb, rows)
    print("  7. Raw Answers")
    write_priority_detail_sheet(wb, orgs, questions)
    print("  8. Priority Detail")
    write_dashboard_sheet(wb, orgs)
    print("  9. Dashboard")
    write_org_report_sheet(wb, orgs)
    print("  10. Organisation Report")

    wb.save(OUTPUT_FILE)
    print(f"\nSaved: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
