"""
Super Filter — Desktop GUI App
================================
Native desktop app (Tkinter) for processing SARI survey Excel files.
Works on macOS and Windows — no extra installs needed beyond Python + openpyxl.

Usage:
    python app.py
"""

import io
import sys
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from pathlib import Path
from collections import defaultdict

import openpyxl

# ═══════════════════════════════════════════════════════════════════════════
# CONFIG
# ═══════════════════════════════════════════════════════════════════════════

RAW_COL_MAP = {
    1: "respondent_id", 2: "submitted_at", 3: "section",
    4: "question_num", 5: "question_id", 6: "question",
    7: "answer", 8: "answer_value", 9: "answer_score",
    10: "max_score", 11: "participant_name", 12: "email",
    13: "job_title", 14: "organisation_name", 15: "organisation_type",
    16: "organisation_size", 17: "stakeholder_category", 18: "pcds_sector",
    19: "district", 20: "role_level", 21: "department",
    22: "age_band", 23: "part_of_group", 24: "parent_company",
}

ALL_OUTPUT_COLUMNS = [
    ("Organisation Name", "organisation_name", "single"),
    ("Parent Company", "parent_company", "single"),
    ("Organisation Type", "organisation_type", "single"),
    ("Organisation Size", "organisation_size", "single"),
    ("Stakeholder Category", "stakeholder_category", "single"),
    ("PDCS Sector", "pcds_sector", "single"),
    ("District", "district", "single"),
    ("Part of Group", "part_of_group", "single"),
    ("Role Level", "role_level", "list"),
    ("Department", "department", "list"),
    ("Age Band", "age_band", "list"),
    ("Job Title", "job_title", "list"),
]

SECTION_ORDER = [
    "Background", "Strategy & Leadership", "Talent & Organisational Culture",
    "Data Management & Readiness", "Infrastructure & Technology",
    "Governance, Policy & Ethics", "Investment", "AI Implementation & Potential Impact",
]

BM_TO_EN_SECTION = {
    "Latar Belakang": "Background", "Strategi & Kepimpinan": "Strategy & Leadership",
    "Bakat & Budaya Organisasi": "Talent & Organisational Culture",
    "Pengurusan Data & Kesiapsiagaan": "Data Management & Readiness",
    "Infrastruktur & Teknologi": "Infrastructure & Technology",
    "Tadbir Urus, Dasar & Etika": "Governance, Policy & Ethics",
    "Pelaburan": "Investment", "Pelaksanaan AI & Impak": "AI Implementation & Potential Impact",
}

ALL_SCORECARD_SECTIONS = [
    "Strategy & Leadership", "Talent & Organisational Culture",
    "Data Management & Readiness", "Infrastructure & Technology",
    "Governance, Policy & Ethics", "Investment", "AI Implementation & Potential Impact",
]

# ═══════════════════════════════════════════════════════════════════════════
# PROCESSING FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════════

def read_raw_data(filepath) -> list[dict]:
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active
    rows = []
    for row_cells in ws.iter_rows(min_row=2, values_only=True):
        record = {}
        for col_idx, key in RAW_COL_MAP.items():
            val = row_cells[col_idx - 1]
            record[key] = str(val).strip() if val is not None else ""
        rows.append(record)
    wb.close()
    return rows


def build_question_order(rows: list[dict]) -> list[tuple]:
    seen = set()
    questions = []
    for row in rows:
        sec = row["section"]
        if sec not in SECTION_ORDER:
            continue
        qid = row["question_id"]
        key = (sec, qid)
        if key not in seen:
            seen.add(key)
            questions.append((sec, qid,
                int(row["question_num"]) if row["question_num"].isdigit() else 0,
                row["question"]))
    questions.sort(key=lambda x: (
        SECTION_ORDER.index(x[0]) if x[0] in SECTION_ORDER else 99, x[2]))
    return questions


def build_org_data(rows: list[dict]) -> dict:
    orgs = defaultdict(lambda: {
        "single": {}, "list": defaultdict(set),
        "answers": defaultdict(set), "scores": defaultdict(list),
    })
    for row in rows:
        org = row["organisation_name"]
        if not org:
            continue
        o = orgs[org]
        for key in ["organisation_name", "parent_company", "organisation_type",
                     "organisation_size", "stakeholder_category", "pcds_sector",
                     "district", "part_of_group"]:
            if key not in o["single"]:
                o["single"][key] = row.get(key, "")
        for key in ["role_level", "department", "age_band", "job_title"]:
            val = row.get(key, "")
            if val:
                o["list"][key].add(val)
        sec = row["section"]
        qid = row["question_id"]
        en_sec = BM_TO_EN_SECTION.get(sec, sec)
        content = row.get("answer", "")
        if content:
            o["answers"][(en_sec, qid)].add(content)
        score_str = row.get("answer_score", "")
        if score_str:
            try:
                o["scores"][(en_sec, qid)].append(float(score_str))
            except ValueError:
                pass
    return orgs


def compute_scores(orgs: dict, questions: list[tuple], scorecard_sections: list,
                   weights: dict = None) -> dict:
    section_qids = defaultdict(list)
    for sec, qid, qnum, qtext in questions:
        if sec in scorecard_sections:
            section_qids[sec].append(qid)

    result = {}
    for org_name, o in orgs.items():
        section_scores = {}
        all_weighted = 0.0
        total_weight = 0.0

        for sec in scorecard_sections:
            sec_total = 0.0
            sec_count = 0
            for qid in section_qids.get(sec, []):
                score_list = o["scores"].get((sec, qid), [])
                if score_list:
                    q_avg = sum(score_list) / len(score_list)
                    w = weights.get(qid, 1.0) if weights else 1.0
                    sec_total += q_avg * w
                    sec_count += w
                    all_weighted += q_avg * w
                    total_weight += w
            section_scores[sec] = round(sec_total / sec_count, 2) if sec_count > 0 else None

        overall = round(all_weighted / total_weight, 2) if total_weight > 0 else None
        result[org_name] = {"section_scores": section_scores, "overall_score": overall}
    return result


def export_to_excel(orgs: dict, questions: list[tuple], scorecard_sections: list,
                    output_columns: list, weights: dict = None) -> bytes:
    wb = openpyxl.Workbook()

    # Pivot
    ws = wb.active
    ws.title = "Pivot"
    org_headers = [c[0] for c in output_columns]
    org_col_count = len(output_columns)
    qcols = [(s, q, t) for s, q, n, t in questions]

    section_spans = []
    cur_sec, sec_start = None, None
    for i, (sec, qid, qtext) in enumerate(qcols):
        col_idx = org_col_count + i + 1
        if sec != cur_sec:
            if cur_sec is not None:
                section_spans.append((cur_sec, sec_start, col_idx - 1))
            cur_sec, sec_start = sec, col_idx
    if cur_sec is not None:
        section_spans.append((cur_sec, sec_start, org_col_count + len(qcols)))

    if org_col_count > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=org_col_count)
        ws.cell(1, 1).value = "Organisation Info"
    for sec_name, sc, ec in section_spans:
        if sc == ec:
            ws.cell(1, sc).value = sec_name
        else:
            ws.merge_cells(start_row=1, start_column=sc, end_row=1, end_column=ec)
            ws.cell(1, sc).value = sec_name

    for ci, h in enumerate(org_headers, 1):
        ws.cell(2, ci).value = h
    for i, (sec, qid, qtext) in enumerate(qcols):
        ws.cell(2, org_col_count + i + 1).value = qid

    row_num = 3
    for org_name in sorted(orgs.keys()):
        o = orgs[org_name]
        list_vals = {}
        for key in ["role_level", "department", "age_band", "job_title"]:
            vals = sorted(o["list"].get(key, set()))
            list_vals[key] = "\n".join(f"{i}. {v}" for i, v in enumerate(vals, 1)) if vals else ""
        for ci, (_, key, agg) in enumerate(output_columns, 1):
            if agg == "single":
                ws.cell(row_num, ci).value = o["single"].get(key, "")
            elif agg == "list":
                ws.cell(row_num, ci).value = list_vals.get(key, "")
        for i, (sec, qid, qtext) in enumerate(qcols):
            answers = o["answers"].get((sec, qid), set())
            ws.cell(row_num, org_col_count + i + 1).value = " | ".join(sorted(answers)) if answers else ""
        row_num += 1

    # Scorecard
    ws2 = wb.create_sheet("Scorecard")
    score_data = compute_scores(orgs, questions, scorecard_sections, weights)
    all_headers = org_headers + scorecard_sections + ["OVERALL"]
    for ci, h in enumerate(all_headers, 1):
        ws2.cell(1, ci).value = h

    row_num = 2
    for org_name in sorted(orgs.keys()):
        o = orgs[org_name]
        sd = score_data.get(org_name, {})
        list_vals = {}
        for key in ["role_level", "department", "age_band", "job_title"]:
            vals = sorted(o["list"].get(key, set()))
            list_vals[key] = "\n".join(f"{i}. {v}" for i, v in enumerate(vals, 1)) if vals else ""
        for ci, (_, key, agg) in enumerate(output_columns, 1):
            if agg == "single":
                ws2.cell(row_num, ci).value = o["single"].get(key, "")
            elif agg == "list":
                ws2.cell(row_num, ci).value = list_vals.get(key, "")
        base = len(output_columns)
        for i, sec in enumerate(scorecard_sections):
            s = sd.get("section_scores", {}).get(sec)
            if s is not None:
                ws2.cell(row_num, base + i + 1).value = s
                ws2.cell(row_num, base + i + 1).number_format = '0.00'
        ov = sd.get("overall_score")
        if ov is not None:
            ws2.cell(row_num, base + len(scorecard_sections) + 1).value = ov
            ws2.cell(row_num, base + len(scorecard_sections) + 1).number_format = '0.00'
        row_num += 1

    # Question Reference
    ws3 = wb.create_sheet("Question Reference")
    ws3.append(["Section", "Question ID", "Question #", "Question Text"])
    for sec, qid, qnum, qtext in questions:
        ws3.append([sec, qid, qnum, qtext])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════════════
# GUI APP
# ═══════════════════════════════════════════════════════════════════════════

class SuperFilterApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Super Filter — SARI Survey Processor")
        self.root.geometry("1400x850")

        # Data
        self.rows = []
        self.questions = []
        self.orgs = {}
        self.filepath = None

        # Column/section toggle vars
        self.col_vars = {}
        self.sec_vars = {}
        self.weight_vars = {}
        self.use_weights = tk.BooleanVar(value=False)

        self._build_ui()

    def _build_ui(self):
        # ── Top bar: file picker ──
        top = ttk.Frame(self.root, padding=5)
        top.pack(fill=tk.X)

        ttk.Label(top, text="File:").pack(side=tk.LEFT, padx=(0, 5))
        self.file_label = ttk.Label(top, text="No file selected", foreground="gray")
        self.file_label.pack(side=tk.LEFT, padx=5)

        ttk.Button(top, text="Open Excel...", command=self._open_file).pack(side=tk.LEFT, padx=5)
        ttk.Button(top, text="Reload", command=self._reload).pack(side=tk.LEFT, padx=5)
        ttk.Button(top, text="Export to Excel", command=self._export).pack(side=tk.RIGHT, padx=5)

        self.status_label = ttk.Label(top, text="")
        self.status_label.pack(side=tk.RIGHT, padx=10)

        # ── Main area: sidebar + tabs ──
        main = ttk.Frame(self.root)
        main.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # Sidebar
        sidebar = ttk.Frame(main, width=280)
        sidebar.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 5))
        sidebar.pack_propagate(False)

        # Column toggles
        col_frame = ttk.LabelFrame(sidebar, text="Org Columns", padding=5)
        col_frame.pack(fill=tk.X, pady=(0, 5))

        col_canvas = tk.Canvas(col_frame, height=200, highlightthickness=0)
        col_scroll = ttk.Scrollbar(col_frame, orient=tk.VERTICAL, command=col_canvas.yview)
        col_inner = ttk.Frame(col_canvas)
        col_inner.bind("<Configure>", lambda e: col_canvas.configure(scrollregion=col_canvas.bbox("all")))
        col_canvas.create_window((0, 0), window=col_inner, anchor="nw")
        col_canvas.configure(yscrollcommand=col_scroll.set)
        col_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        col_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        for label, key, agg in ALL_OUTPUT_COLUMNS:
            var = tk.BooleanVar(value=True)
            self.col_vars[key] = var
            ttk.Checkbutton(col_inner, text=label, variable=var).pack(anchor=tk.W, pady=1)

        # Section toggles
        sec_frame = ttk.LabelFrame(sidebar, text="Scorecard Sections", padding=5)
        sec_frame.pack(fill=tk.X, pady=(0, 5))

        for sec in ALL_SCORECARD_SECTIONS:
            var = tk.BooleanVar(value=True)
            self.sec_vars[sec] = var
            ttk.Checkbutton(sec_frame, text=sec, variable=var).pack(anchor=tk.W, pady=1)

        # Weightage toggle
        weight_frame = ttk.LabelFrame(sidebar, text="Weightage", padding=5)
        weight_frame.pack(fill=tk.X, pady=(0, 5))
        ttk.Checkbutton(weight_frame, text="Enable custom weightage",
                        variable=self.use_weights, command=self._toggle_weights).pack(anchor=tk.W)

        self.weight_inner = ttk.Frame(weight_frame)
        self.weight_inner.pack(fill=tk.X, pady=(5, 0))

        # Search
        search_frame = ttk.LabelFrame(sidebar, text="Filter", padding=5)
        search_frame.pack(fill=tk.X)
        self.search_var = tk.StringVar()
        self.search_var.trace_add("write", lambda *a: self._refresh_tables())
        ttk.Entry(search_frame, textvariable=self.search_var).pack(fill=tk.X)

        # ── Tabs ──
        self.notebook = ttk.Notebook(main)
        self.notebook.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Pivot tab
        self.pivot_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.pivot_frame, text="Pivot Table")
        self._build_table(self.pivot_frame, "pivot")

        # Scorecard tab
        self.sc_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.sc_frame, text="Scorecard")
        self._build_table(self.sc_frame, "scorecard")

        # Question Reference tab
        self.qr_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.qr_frame, text="Question Reference")
        self._build_table(self.qr_frame, "qref")

    def _build_table(self, parent, name):
        """Build a Treeview with scrollbars inside a frame."""
        frame = ttk.Frame(parent)
        frame.pack(fill=tk.BOTH, expand=True)

        cols = ("dummy",)  # placeholder, rebuilt on data load
        tree = ttk.Treeview(frame, columns=cols, show="headings", height=20)
        tree.heading("dummy", text="Load a file to see data")

        vsb = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=tree.yview)
        hsb = ttk.Scrollbar(frame, orient=tk.HORIZONTAL, command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        frame.grid_rowconfigure(0, weight=1)
        frame.grid_columnconfigure(0, weight=1)

        setattr(self, f"tree_{name}", tree)

    def _open_file(self):
        path = filedialog.askopenfilename(
            title="Select SARI Survey Excel File",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        if path:
            self.filepath = path
            self.file_label.config(text=Path(path).name, foreground="black")
            self._process()

    def _reload(self):
        if self.filepath:
            self._process()

    def _process(self):
        if not self.filepath:
            return
        try:
            self.status_label.config(text="Processing...")
            self.root.update()

            self.rows = read_raw_data(self.filepath)
            self.questions = build_question_order(self.rows)
            self.orgs = build_org_data(self.rows)

            self._build_weight_inputs()
            self._refresh_tables()

            n_orgs = len(self.orgs)
            n_qs = len(self.questions)
            self.status_label.config(
                text=f"Loaded: {n_orgs} orgs, {n_qs} questions")
        except Exception as e:
            messagebox.showerror("Error", str(e))
            self.status_label.config(text="Error loading file")

    def _build_weight_inputs(self):
        """Rebuild weight input fields based on loaded questions."""
        for w in self.weight_inner.winfo_children():
            w.destroy()

        self.weight_vars = {}
        # Only show scored questions
        scored = [(s, q, n, t) for s, q, n, t in self.questions
                  if s in ALL_SCORECARD_SECTIONS]

        if not scored:
            ttk.Label(self.weight_inner, text="No scored questions loaded").pack()
            return

        wcanvas = tk.Canvas(self.weight_inner, height=200, highlightthickness=0)
        wscroll = ttk.Scrollbar(self.weight_inner, orient=tk.VERTICAL, command=wcanvas.yview)
        winner = ttk.Frame(wcanvas)
        winner.bind("<Configure>", lambda e: wcanvas.configure(scrollregion=wcanvas.bbox("all")))
        wcanvas.create_window((0, 0), window=winner, anchor="nw")
        wcanvas.configure(yscrollcommand=wscroll.set)
        wcanvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        wscroll.pack(side=tk.RIGHT, fill=tk.Y)

        for sec, qid, qnum, qtext in scored:
            row = ttk.Frame(winner)
            row.pack(fill=tk.X, pady=1)
            ttk.Label(row, text=f"{qid}", width=18, anchor=tk.W).pack(side=tk.LEFT)
            var = tk.DoubleVar(value=1.0)
            self.weight_vars[qid] = var
            ttk.Spinbox(row, from_=0, to=10, increment=0.5, textvariable=var,
                        width=5).pack(side=tk.LEFT)

    def _toggle_weights(self):
        self._refresh_tables()

    def _get_selected_cols(self):
        return [(l, k, a) for l, k, a in ALL_OUTPUT_COLUMNS if self.col_vars[k].get()]

    def _get_selected_sections(self):
        return [s for s in ALL_SCORECARD_SECTIONS if self.sec_vars[s].get()]

    def _get_weights(self):
        if not self.use_weights.get():
            return None
        return {qid: var.get() for qid, var in self.weight_vars.items()}

    def _refresh_tables(self):
        if not self.orgs:
            return
        self._refresh_pivot()
        self._refresh_scorecard()
        self._refresh_qref()

    def _refresh_pivot(self):
        tree = self.tree_pivot
        tree.delete(*tree.get_children())

        org_cols = self._get_selected_cols()
        qcols = [(s, q, t) for s, q, n, t in self.questions]
        all_headers = [c[0] for c in org_cols] + [qid for _, qid, _ in qcols]

        tree["columns"] = all_headers
        for i, h in enumerate(all_headers):
            tree.heading(h, text=h)
            tree.column(h, width=120, minwidth=60, anchor=tk.W)

        search = self.search_var.get().lower()
        for org_name in sorted(self.orgs.keys()):
            if search and search not in org_name.lower():
                continue
            o = self.orgs[org_name]
            list_vals = {}
            for key in ["role_level", "department", "age_band", "job_title"]:
                vals = sorted(o["list"].get(key, set()))
                list_vals[key] = "\n".join(f"{i}. {v}" for i, v in enumerate(vals, 1)) if vals else ""

            row = []
            for _, key, agg in org_cols:
                if agg == "single":
                    row.append(o["single"].get(key, ""))
                elif agg == "list":
                    row.append(list_vals.get(key, ""))

            for sec, qid, qtext in qcols:
                answers = o["answers"].get((sec, qid), set())
                row.append(" | ".join(sorted(answers)) if answers else "")

            tree.insert("", tk.END, values=row)

    def _refresh_scorecard(self):
        tree = self.tree_scorecard
        tree.delete(*tree.get_children())

        org_cols = self._get_selected_cols()
        sections = self._get_selected_sections()
        weights = self._get_weights()
        score_data = compute_scores(self.orgs, self.questions, sections, weights)

        all_headers = [c[0] for c in org_cols] + sections + ["OVERALL"]
        tree["columns"] = all_headers
        for i, h in enumerate(all_headers):
            tree.heading(h, text=h)
            tree.column(h, width=100, minwidth=60, anchor=tk.W)

        search = self.search_var.get().lower()
        for org_name in sorted(self.orgs.keys()):
            if search and search not in org_name.lower():
                continue
            o = self.orgs[org_name]
            sd = score_data.get(org_name, {})
            list_vals = {}
            for key in ["role_level", "department", "age_band", "job_title"]:
                vals = sorted(o["list"].get(key, set()))
                list_vals[key] = "\n".join(f"{i}. {v}" for i, v in enumerate(vals, 1)) if vals else ""

            row = []
            for _, key, agg in org_cols:
                if agg == "single":
                    row.append(o["single"].get(key, ""))
                elif agg == "list":
                    row.append(list_vals.get(key, ""))

            for sec in sections:
                s = sd.get("section_scores", {}).get(sec)
                row.append(f"{s:.2f}" if s is not None else "-")

            ov = sd.get("overall_score")
            row.append(f"{ov:.2f}" if ov is not None else "-")
            tree.insert("", tk.END, values=row)

    def _refresh_qref(self):
        tree = self.tree_qref
        tree.delete(*tree.get_children())

        tree["columns"] = ("Section", "Question ID", "Question #", "Question Text")
        for h in tree["columns"]:
            tree.heading(h, text=h)
        tree.column("Section", width=200)
        tree.column("Question ID", width=120)
        tree.column("Question #", width=80)
        tree.column("Question Text", width=500)

        for sec, qid, qnum, qtext in self.questions:
            tree.insert("", tk.END, values=(sec, qid, qnum, qtext))

    def _export(self):
        if not self.orgs:
            messagebox.showwarning("No Data", "Load a file first.")
            return

        path = filedialog.asksaveasfilename(
            title="Save Processed Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")])
        if not path:
            return

        try:
            excel_bytes = export_to_excel(
                self.orgs, self.questions,
                self._get_selected_sections(),
                self._get_selected_cols(),
                self._get_weights(),
            )
            with open(path, "wb") as f:
                f.write(excel_bytes)
            messagebox.showinfo("Success", f"Saved to:\n{path}")
        except Exception as e:
            messagebox.showerror("Export Error", str(e))


def main():
    root = tk.Tk()

    # macOS: bring app to front
    try:
        root.lift()
        root.attributes("-topmost", True)
        root.after(100, lambda: root.attributes("-topmost", False))
    except Exception:
        pass

    app = SuperFilterApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
